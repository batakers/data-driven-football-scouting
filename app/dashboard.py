import os
import sys
import html
import json
import re
import io
import unicodedata
from datetime import date
from pathlib import Path
import joblib

import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# Fix for ModuleNotFoundError: No module named 'src'
# Ensures the project root is in the path when running from the app/ folder
PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import importlib
import src.similarity
import src.scouting_rationale
importlib.reload(src.similarity)
importlib.reload(src.scouting_rationale)
from src.similarity import PlayerSimilarity
from src.role_mapping import COMPATIBLE_ROLE_THRESHOLD
from src.scouting_rationale import build_scouting_signals, build_scout_checks, build_rationale_summary, compute_confidence_level
from src.league_tiers import (
    enrich_dataframe as enrich_league_tier,
    get_tier_short,
    get_scout_note,
    TIER_SHORT_LABELS,
)
from src.enrich_contract import (
    enrich_contract,
    get_contract_scout_note,
    get_contract_badge,
    CONTRACT_STATUS_LABELS,
)

# =========================
# Optional dependency checks
# =========================
try:
    import openpyxl  # noqa: F401
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

try:
    from fpdf import FPDF
    _FPDF_AVAILABLE = True
except ImportError:
    _FPDF_AVAILABLE = False


# =========================
# Page Configuration
# =========================
st.set_page_config(
    page_title="Football Talent Scouting Dashboard",
    page_icon="⚽",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# =========================
# Theme Styling
# Uses Streamlit native CSS variables so it adapts to light/dark mode.
# =========================
PLOTLY_TEMPLATE = "plotly_dark"  # Keep charts dark for professional look
REFERENCE_LINE_COLOR = "#94a3b8"

st.markdown(
    """
<style>
.app-hero {
    margin: 0.4rem 0 1.4rem 0;
}

.app-title {
    color: var(--text-color);
    font-size: clamp(2rem, 3vw, 3.2rem);
    font-weight: 900;
    letter-spacing: -0.04em;
    line-height: 1.05;
}

.app-subtitle {
    color: var(--text-color);
    opacity: 0.72;
    font-size: 1rem;
    margin-top: 0.55rem;
    max-width: 900px;
}

.kpi-card {
    background-color: var(--secondary-background-color);
    border: 1px solid rgba(128, 128, 128, 0.25);
    border-radius: 14px;
    padding: 18px 20px;
    box-shadow: 0 4px 14px rgba(0, 0, 0, 0.10);
    min-height: 105px;
    margin-bottom: 25px;
}

.kpi-label {
    color: var(--text-color);
    opacity: 0.70;
    font-size: 0.85rem;
    font-weight: 600;
    margin-bottom: 10px;
}

.kpi-value {
    color: var(--text-color);
    font-size: 2rem;
    font-weight: 800;
    line-height: 1.15;
}

.methodology-box {
    background-color: rgba(59, 130, 246, 0.12);
    border: 1px solid rgba(59, 130, 246, 0.35);
    color: var(--text-color);
    padding: 0.85rem 1rem;
    border-radius: 8px;
    font-size: 0.92rem;
    margin-top: 15px;
    margin-bottom: 10px;
}

.disclaimer-box {
    background-color: rgba(245, 158, 11, 0.12);
    border: 1px solid rgba(245, 158, 11, 0.45);
    color: var(--text-color);
    padding: 0.85rem 1rem;
    border-radius: 8px;
    font-size: 0.92rem;
    margin-top: 15px;
}

div[data-testid="stExpander"] {
    border-radius: 10px;
}
.advanced-stats-badge {
    background-color: #059669;
    color: white;
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 0.75rem;
    font-weight: 600;
    margin-left: 10px;
    display: inline-block;
}
.basic-stats-badge {
    background-color: #4b5563;
    color: white;
    padding: 2px 8px;
    border-radius: 4px;
    font-size: 0.75rem;
    font-weight: 600;
    margin-left: 10px;
    display: inline-block;
}
/* ── League Tier Badges ─────────────────────────────────────────────── */
.tier-badge {
    padding: 2px 9px;
    border-radius: 4px;
    font-size: 0.75rem;
    font-weight: 700;
    display: inline-block;
    letter-spacing: 0.02em;
}
.tier-1 { background-color: #1d4ed8; color: #fff; }   /* Elite  -  blue */
.tier-2 { background-color: #0f766e; color: #fff; }   /* Strong  -  teal */
.tier-3 { background-color: #b45309; color: #fff; }   /* Competitive  -  amber */
.tier-4 { background-color: #6b7280; color: #fff; }   /* Developing  -  grey */

/* ── Contract Status Badges ─────────────────────────────────────────── */
.contract-badge {
    padding: 2px 9px;
    border-radius: 4px;
    font-size: 0.75rem;
    font-weight: 700;
    display: inline-block;
    letter-spacing: 0.02em;
    margin-left: 6px;
}
.contract-expiring { background-color: #dc2626; color: #fff; }  /* red  -  urgent */
.contract-short    { background-color: #ea580c; color: #fff; }  /* orange */
.contract-medium   { background-color: #ca8a04; color: #fff; }  /* yellow-dark */
.contract-long     { background-color: #16a34a; color: #fff; }  /* green */
.contract-expired  { background-color: #7c3aed; color: #fff; }  /* purple */
.contract-unknown  { background-color: #6b7280; color: #fff; }  /* grey */
.target-panel {
    background: linear-gradient(135deg, rgba(37, 99, 235, 0.10), rgba(15, 23, 42, 0.02));
    border: 1px solid rgba(37, 99, 235, 0.24);
    border-radius: 16px;
    padding: 18px 20px;
    margin: 8px 0 16px 0;
}
.target-name {
    color: var(--text-color);
    font-size: 1.45rem;
    font-weight: 800;
    margin-bottom: 5px;
}
.target-summary {
    color: var(--text-color);
    opacity: 0.80;
    font-size: 0.96rem;
    margin-bottom: 10px;
}
.target-context {
    color: var(--text-color);
    opacity: 0.72;
    font-size: 0.88rem;
}

/* ── Player Hero (compact header with name + badges) ────────────────── */
.player-hero {
    background: linear-gradient(135deg, rgba(37, 99, 235, 0.08), rgba(15, 23, 42, 0.02));
    border: 1px solid rgba(37, 99, 235, 0.22);
    border-radius: 14px;
    padding: 14px 18px;
    margin: 8px 0 14px 0;
    display: flex;
    flex-wrap: wrap;
    align-items: baseline;
    gap: 10px 14px;
}
.player-hero .hero-name {
    color: var(--text-color);
    font-size: 1.55rem;
    font-weight: 800;
    line-height: 1.1;
    margin-right: 6px;
}
.player-hero .hero-meta {
    color: var(--text-color);
    opacity: 0.78;
    font-size: 0.95rem;
    font-weight: 500;
}
.player-hero .hero-meta .dot {
    opacity: 0.5;
    margin: 0 4px;
}

/* ── Info Line (compact row: chips + caption, no cards) ─────────────── */
.info-line {
    display: flex;
    flex-wrap: wrap;
    gap: 6px 10px;
    align-items: center;
    color: var(--text-color);
    font-size: 0.92rem;
    padding: 6px 0 12px 0;
    opacity: 0.88;
}
.info-line .info-chip {
    border: 1px solid rgba(100, 116, 139, 0.28);
    background: rgba(148, 163, 184, 0.10);
    border-radius: 999px;
    padding: 3px 10px;
    font-size: 0.86rem;
    font-weight: 600;
    display: inline-flex;
    align-items: center;
    gap: 4px;
}
.info-line .info-chip .chip-label {
    opacity: 0.65;
    font-weight: 500;
}

/* ── Form Strip (single compact row for form trend) ─────────────────── */
.form-strip {
    display: flex;
    flex-wrap: wrap;
    gap: 18px;
    align-items: center;
    padding: 10px 14px;
    background: rgba(148, 163, 184, 0.08);
    border: 1px solid rgba(100, 116, 139, 0.20);
    border-radius: 10px;
    margin: 6px 0 14px 0;
    font-size: 0.92rem;
}
.form-strip .form-label {
    font-weight: 700;
    color: var(--text-color);
}
.form-strip .form-metric {
    color: var(--text-color);
    opacity: 0.82;
}
.form-strip .form-metric b {
    opacity: 1;
    font-weight: 700;
}
.form-strip.trend-rising    { border-left: 4px solid #16a34a; }
.form-strip.trend-stable    { border-left: 4px solid #ca8a04; }
.form-strip.trend-declining { border-left: 4px solid #dc2626; }
.form-strip.trend-unknown   { border-left: 4px solid #6b7280; }
.chip-row {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin: 6px 0 14px 0;
}
.profile-chip {
    border: 1px solid rgba(100, 116, 139, 0.28);
    background: rgba(148, 163, 184, 0.12);
    color: var(--text-color);
    border-radius: 999px;
    padding: 5px 10px;
    font-size: 0.84rem;
    font-weight: 600;
}
.data-chip {
    border: 1px solid rgba(37, 99, 235, 0.30);
    background: rgba(59, 130, 246, 0.10);
}
</style>
""",
    unsafe_allow_html=True,
)

def kpi_card(label: str, value: str) -> None:
    """Render a custom KPI card that works better in both light and dark mode."""
    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{value}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_player_hero(
    name: str,
    badges_html: str = "",
    meta_items: list[str] | None = None,
) -> None:
    """Render a compact player hero strip with name, badges, and inline metadata.

    Replaces 4-8 redundant KPI cards with a single dense header band.
    """
    meta_parts = []
    for i, item in enumerate(meta_items or []):
        if not item:
            continue
        if i > 0:
            meta_parts.append('<span class="dot">·</span>')
        meta_parts.append(html.escape(str(item)))
    meta_html = "".join(meta_parts)
    st.markdown(
        f"""
        <div class="player-hero">
            <div class="hero-name">{html.escape(str(name))}</div>
            {badges_html}
            <div class="hero-meta">{meta_html}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_info_line(pairs: list[tuple[str, str]]) -> None:
    """Render an inline chip row where each chip shows a compact label + value.

    Use this for secondary context info (Position, Age, Club) that doesn't
    warrant dedicated KPI cards.
    """
    chips = []
    for label, value in pairs:
        if value is None or (isinstance(value, str) and not value.strip()):
            continue
        chips.append(
            f'<span class="info-chip">'
            f'<span class="chip-label">{html.escape(str(label))}:</span>'
            f' {html.escape(str(value))}'
            f'</span>'
        )
    if chips:
        st.markdown(
            f'<div class="info-line">{"".join(chips)}</div>',
            unsafe_allow_html=True,
        )

def render_form_strip(
    form_val: str,
    goals_delta,
    assists_delta,
    mins_ratio,
) -> None:
    """Render a single-row form trend strip to replace 4 redundant KPI cards."""
    trend_class_map = {
        "Rising":    "trend-rising",
        "Stable":    "trend-stable",
        "Declining": "trend-declining",
    }
    trend_class = trend_class_map.get(str(form_val), "trend-unknown")
    form_icon = {
        "Rising": "↑",
        "Stable": "→",
        "Declining": "↓",
        "Insufficient Data": " - ",
    }.get(str(form_val), " - ")

    def _fmt_signed(v, fmt=".2f", suffix=""):
        if v is None:
            return "N/A"
        try:
            fv = float(v)
            if pd.isna(fv):
                return "N/A"
            return f"{fv:+{fmt}}{suffix}"
        except (TypeError, ValueError):
            return "N/A"

    def _fmt_ratio(v):
        if v is None:
            return "N/A"
        try:
            fv = float(v)
            if pd.isna(fv):
                return "N/A"
            return f"{fv:.2f}x"
        except (TypeError, ValueError):
            return "N/A"

    st.markdown(
        f"""
        <div class="form-strip {trend_class}">
            <div class="form-label">{form_icon} Form: {html.escape(str(form_val))}</div>
            <div class="form-metric">Goals/90 <b>{_fmt_signed(goals_delta)}</b></div>
            <div class="form-metric">Assists/90 <b>{_fmt_signed(assists_delta)}</b></div>
            <div class="form-metric">Minutes <b>{_fmt_ratio(mins_ratio)}</b></div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.caption("Form compares the last 6 months to the previous 6 months · requires ≥450 minutes per window.")

def methodology_box(message: str) -> None:
    st.markdown(
        f"""
        <div class="methodology-box">
            💡 <strong>Methodology:</strong> {message}
        </div>
        """,
        unsafe_allow_html=True,
    )

def disclaimer_box(message: str) -> None:
    st.markdown(
        f"""
        <div class="disclaimer-box">
            ⚠️ <strong>Disclaimer:</strong> {message}
        </div>
        """,
        unsafe_allow_html=True,
    )

def league_tier_badge(competition_id: str | None, label: str | None = None) -> str:
    """Return an HTML badge string for a league tier. Safe to embed in st.markdown."""
    tier = get_tier_short(competition_id)
    tier_num = {"Elite": 1, "Strong": 2, "Competitive": 3, "Developing": 4}.get(tier, 4)
    display = html.escape(label or tier)
    return f'<span class="tier-badge tier-{tier_num}">{display}</span>'

def contract_status_badge(status: str | None) -> str:
    """Return an HTML badge string for a contract status. Safe to embed in st.markdown."""
    if not status or pd.isna(status):
        status = "Unknown"
    css_map = {
        "Expiring": "contract-expiring",
        "Short":    "contract-short",
        "Medium":   "contract-medium",
        "Long":     "contract-long",
        "Expired":  "contract-expired",
        "Unknown":  "contract-unknown",
    }
    css_class = css_map.get(status, "contract-unknown")
    badge_text = get_contract_badge(status)
    return f'<span class="contract-badge {css_class}">{html.escape(badge_text)}</span>'

def scalar_bool(value) -> bool:
    if pd.isna(value):
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes"}

def _safe_float(value, default: float = 0.0) -> float:
    """Safely convert a value to float, returning default on failure."""
    try:
        v = float(value)
        import math
        return default if math.isnan(v) else v
    except (TypeError, ValueError):
        return default

def format_stats_season(value) -> str:
    if pd.isna(value):
        return "N/A"
    raw = str(value).strip()
    try:
        raw = str(int(float(raw)))
    except ValueError:
        raw = raw.replace(".0", "")
    if len(raw) != 4:
        return "N/A"
    return f"20{raw[:2]}/20{raw[2:]}"

def format_role_tags(value) -> str:
    if pd.isna(value):
        return "N/A"
    tags = [tag.strip() for tag in str(value).split(",") if tag.strip()]
    return ", ".join(tags) if tags else "N/A"

def format_foot_profile(value, has_metadata: bool) -> str:
    if not has_metadata or pd.isna(value):
        return "Foot unknown"
    foot = str(value).strip().lower()
    if foot in {"left", "right"}:
        return f"{foot.title()}-footed"
    if foot == "both":
        return "Two-footed"
    return "Foot unknown"

def role_tags_are_redundant(primary_role: str, role_tags: str) -> bool:
    if not role_tags or role_tags == "N/A":
        return True
    tags = [tag.strip() for tag in role_tags.split(",") if tag.strip()]
    return len(tags) == 1 and tags[0].upper() == str(primary_role).upper()

def parse_ui_role_tags(value) -> set[str]:
    if pd.isna(value):
        return set()
    return {
        tag.strip().upper()
        for tag in str(value).replace("/", ",").split(",")
        if tag.strip()
    }

def scout_role_fit_label(primary_role, role_tags, matched_as, raw_fit_type) -> str:
    """Translate technical role-fit metadata into scout-friendly labels."""
    tags = parse_ui_role_tags(role_tags)
    matched = "" if pd.isna(matched_as) else str(matched_as).strip().upper()
    raw = "" if pd.isna(raw_fit_type) else str(raw_fit_type).strip()

    if tags & {"GK"}:
        return "Goalkeeper Role Match" if matched == "GK" else "Not a Tactical Fit"

    defensive_central = {"CDM", "CM", "CB"}
    wide_attack_mid = {"LW", "RW", "LM", "RM", "CAM", "CM"}
    front_line = {"ST", "CF", "LW", "RW", "CAM"}
    defensive_utility = {"LB", "CB", "RB", "LWB", "RWB"}
    wide_side = {"LW", "RW", "LM", "RM", "LB", "RB", "LWB", "RWB"}

    if {"CDM", "CB"}.issubset(tags) or len(tags & defensive_central) >= 3:
        return "Defensive Central Hybrid"
    if len(tags & wide_attack_mid) >= 4 and bool(tags & {"CAM", "CM"}):
        return "Creative Utility Player"
    if len(tags & front_line) >= 3 and bool(tags & {"ST", "CF"}):
        return "Front-Line Hybrid"
    if len(tags & {"LB", "CB", "RB"}) >= 2:
        return "Defensive Utility Player"
    if len(tags & {"LW", "RW", "LM", "RM"}) >= 3:
        return "Wide-Side Specialist"
    if len(tags & wide_side) >= 3:
        return "Wide-Side Specialist"

    mapping = {
        "Primary Role Match": "Natural Role Match",
        "Secondary Role Match": "Proven Secondary Fit",
        "Shared Role Tag": "Proven Multi-Role Fit",
        "Compatible Role": "Tactical Adjacent Fit",
        "Broad Position": "Broad Position Fit",
        "No Tactical Fit": "Not a Tactical Fit",
        "No Exact Role Fit": "Not an Exact Role Fit",
        "Below Role Threshold": "Below Tactical Fit Threshold",
        "No Role Metadata": "Broad Position Fallback",
    }
    return mapping.get(raw, raw or "Role Fit Unknown")

def render_chip_row(labels: list[str], extra_class: str = "") -> str:
    class_name = f"profile-chip {extra_class}".strip()
    chips = "".join(
        f'<span class="{class_name}">{html.escape(str(label))}</span>'
        for label in labels
        if label and str(label).strip()
    )
    return f'<div class="chip-row">{chips}</div>'

def format_score(value) -> str:
    if pd.isna(value):
        return "N/A"
    return f"{float(value) * 100:.2f}%"

def price_status_label(value_diff, target_value) -> str:
    if pd.isna(value_diff) or pd.isna(target_value) or float(target_value) <= 0:
        return "Unknown"
    threshold = abs(float(target_value)) * 0.15
    diff = float(value_diff)
    if diff < -threshold:
        return "Lower Cost"
    if diff > threshold:
        return "Higher Cost"
    return "Similar Cost"

def format_match_method(value) -> str:
    if pd.isna(value):
        return "N/A"
    method = str(value).strip()
    if not method:
        return "N/A"
    return method.replace("_", " ").title()

def format_match_confidence(value, fallback: str = "N/A") -> str:
    if pd.isna(value):
        return fallback
    return f"{float(value):.0f}%"

def fuzzy_confidence_note(method, score) -> str | None:
    method_raw = "" if pd.isna(method) else str(method).strip().lower()
    if not method_raw.startswith("fuzzy"):
        return None
    try:
        score_value = float(score)
    except (TypeError, ValueError):
        score_value = None
    if score_value is not None and score_value >= 99.5:
        return "100% is the fuzzy text-match score after validation gates; it does not mean the match method was exact."
    return "Fuzzy matches are accepted only after validation gates such as role, team, season, and duplicate-source checks."

def format_currency(value) -> str:
    if pd.isna(value):
        return "N/A"
    return f"€{float(value):,.0f}"

def format_percentage(value, decimals: int = 1) -> str:
    if pd.isna(value):
        return "N/A"
    return f"{float(value) * 100:.{decimals}f}%"

LEAGUE_LABELS = {
    "GB1": "Premier League",
    "ES1": "La Liga",
    "L1": "Bundesliga",
    "IT1": "Serie A",
    "FR1": "Ligue 1",
    "NL1": "Eredivisie",
    "PO1": "Liga Portugal",
    "BE1": "Jupiler Pro League",
    "TR1": "Super Lig",
    "SC1": "Scottish Premiership",
    "DK1": "Superligaen",
    "RU1": "Russian Premier League",
    "UKR1": "Ukrainian Premier League",
    "GR1": "Super League Greece",
    "SA1": "Saudi Pro League",
    "BRA1": "Brasileirao Serie A",
    "NO1": "Eliteserien",
    "MLS1": "MLS",
    "SER1": "Serbian SuperLiga",
    "RO1": "Romanian SuperLiga",
    "PL1": "Ekstraklasa",
    "SE1": "Allsvenskan",
    "A1": "Austrian Bundesliga",
    "ARG1": "Liga Profesional",
    "C1": "Championship",
    "AUS1": "A-League Men",
    "MEX1": "Liga MX",
    "JAP1": "J1 League",
    "KR1": "K League 1",
}

CLUB_SHORT_OVERRIDES = {
    "Association sportive de Monaco Football Club": "AS Monaco",
    "Brighton and Hove Albion Football Club": "Brighton",
    "Club Atletico de Madrid S.A.D.": "Atletico Madrid",
    "Futebol Clube do Porto": "FC Porto",
    "Futbol Club Barcelona": "Barcelona",
    "Olympique Gymnaste Club Nice Cote d'Azur": "OGC Nice",
    "Olympique Gymnaste Club Nice Côte d'Azur": "OGC Nice",
    "Reial Club Deportiu Espanyol de Barcelona S.A.D.": "Espanyol",
    "Real Madrid Club de Futbol": "Real Madrid",
    "Real Madrid Club de Fútbol": "Real Madrid",
    "The Celtic Football Club": "Celtic FC",
}

CLUB_NAME_SUFFIXES = [
    " Football Club",
    " Fútbol Club S.A.D.",
    " Fútbol Club S. A. D.",
    " Futbol Club S.A.D.",
    " Futbol Club S. A. D.",
    " Fútbol Club",
    " Futbol Club",
    " Club de Fútbol",
    " Club de Futbol",
    " S.A.D.",
    " S. A. D.",
]

def current_league_label(row: pd.Series) -> str:
    comp_id = row.get("current_club_domestic_competition_id", None)
    if pd.notna(comp_id) and str(comp_id).strip():
        comp = str(comp_id).strip()
        return LEAGUE_LABELS.get(comp, comp)

    for col in ["current_league", "target_league_x", "target_league_y", "target_league"]:
        if col in row and pd.notna(row[col]) and str(row[col]).strip():
            value = str(row[col]).strip()
            return value.split("-", 1)[1] if "-" in value else value

    return "N/A"

def current_club_label(row: pd.Series) -> str:
    for col in ["current_club_name", "current_club", "club"]:
        if col in row and pd.notna(row[col]) and str(row[col]).strip():
            return str(row[col]).strip()
    return "N/A"

def comparable_id(value) -> str:
    if pd.isna(value):
        return ""
    raw = str(value).strip()
    try:
        return str(int(float(raw)))
    except ValueError:
        return raw

def matching_current_club_id(row: pd.Series) -> bool:
    current_id = comparable_id(row.get("current_club_id", None))
    club_id = comparable_id(row.get("club_id", None))
    return bool(current_id and club_id and current_id == club_id)

def clean_club_display_name(club_name: str) -> str:
    if club_name == "N/A":
        return club_name

    cleaned = CLUB_SHORT_OVERRIDES.get(club_name, club_name)
    for suffix in CLUB_NAME_SUFFIXES:
        if cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)].strip()
            break

    if cleaned.startswith("The "):
        cleaned = cleaned[4:].strip()

    if len(cleaned) > 36:
        cleaned = f"{cleaned[:33].rstrip()}..."
    return cleaned

def current_club_display_label(row: pd.Series) -> str:
    full_club = current_club_label(row)
    short_club = row.get("club", None)
    if (
        full_club != "N/A"
        and pd.notna(short_club)
        and str(short_club).strip()
        and matching_current_club_id(row)
        and len(str(short_club).strip()) < len(full_club)
    ):
        return str(short_club).strip()
    return clean_club_display_name(full_club)

def club_league_label(row: pd.Series) -> str:
    club = current_club_display_label(row)
    league = current_league_label(row)
    if club == "N/A":
        return league
    if league == "N/A":
        return club
    return f"{club} · {league}"

def add_download_context_columns(df: pd.DataFrame) -> pd.DataFrame:
    download_df = df.copy()
    download_df["current_club"] = download_df.apply(current_club_label, axis=1)
    download_df["current_league"] = download_df.apply(current_league_label, axis=1)
    return download_df

def safe_filename_label(value, fallback: str = "player") -> str:
    label = str(value or fallback).strip()
    label = re.sub(r"[^A-Za-z0-9_-]+", "_", label)
    return label.strip("_") or fallback


# =========================
# Phase 12  -  Export helpers
# =========================

def _ascii_safe(text: str) -> str:
    """Normalize unicode text to ASCII-safe string for PDF output."""
    return unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")


def build_excel_bytes(
    data_df: pd.DataFrame,
    sheet_name: str = "Shortlist",
    metadata_rows: list[dict] | None = None,
) -> bytes:
    """
    Build a multi-sheet Excel file in memory and return raw bytes.
    Requires openpyxl. Returns None if unavailable.
    """
    if not _OPENPYXL_AVAILABLE:
        return None
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        data_df.to_excel(writer, sheet_name=sheet_name, index=False)
        if metadata_rows:
            meta_df = pd.DataFrame(metadata_rows)
            meta_df.to_excel(writer, sheet_name="Metadata", index=False)
    return buf.getvalue()


def build_shortlist_excel(
    df: pd.DataFrame,
    active_filters: dict | None = None,
) -> bytes | None:
    """
    Build the full shortlist Excel export with all Phase 8-11 columns and a metadata sheet.
    """
    if not _OPENPYXL_AVAILABLE:
        return None

    export_df = add_download_context_columns(df).copy()

    # Ensure Phase 8-11 columns are present
    phase_cols = [
        "league_tier_short",
        "contract_status_label",
        "contract_months_remaining",
        "form_trend",
        "form_trend_goals",
        "form_trend_assists",
    ]
    for col in phase_cols:
        if col not in export_df.columns:
            export_df[col] = None

    # Build metadata sheet
    today_str = date.today().isoformat()
    filter_summary = "; ".join(
        f"{k}={v}" for k, v in (active_filters or {}).items()
    ) or "None"
    metadata_rows = [
        {"Field": "Export Date", "Value": today_str},
        {"Field": "Active Filters", "Value": filter_summary},
        {"Field": "Total Players", "Value": str(len(export_df))},
        {
            "Field": "Methodology Note",
            "Value": (
                "Players are ranked by value gap (predicted value / current market value). "
                "League tier, contract status, and form trend are enrichment layers from Phases 8-11. "
                "This export is a statistical scouting lead  -  validate with professional human scouting."
            ),
        },
    ]
    return build_excel_bytes(export_df, sheet_name="Shortlist", metadata_rows=metadata_rows)


def build_similarity_excel(
    df: pd.DataFrame,
    target_name: str,
) -> bytes | None:
    """Build the similarity comparison Excel export."""
    if not _OPENPYXL_AVAILABLE:
        return None
    export_df = add_download_context_columns(df).copy()
    today_str = date.today().isoformat()
    metadata_rows = [
        {"Field": "Export Date", "Value": today_str},
        {"Field": "Target Player", "Value": target_name},
        {"Field": "Total Alternatives", "Value": str(len(export_df))},
        {
            "Field": "Methodology Note",
            "Value": (
                "Similarity scores combine statistical profile match, role compatibility, and foot-side fit. "
                "Market value is used for post-match filtering only, not for similarity calculation."
            ),
        },
    ]
    return build_excel_bytes(export_df, sheet_name="Alternatives", metadata_rows=metadata_rows)


def build_scouting_pdf(
    player_name: str,
    row: pd.Series,
    rationale_text: str,
    signals_df: pd.DataFrame,
    scout_checks: list[str],
) -> bytes | None:
    """
    Generate a clean, printable PDF scouting brief for a single player.
    Returns raw bytes or None if fpdf2 is unavailable.
    """
    if not _FPDF_AVAILABLE:
        return None

    def s(text) -> str:
        """ASCII-safe string for PDF."""
        return _ascii_safe(str(text) if text is not None else "")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_margins(18, 15, 18)

    # ── Header ──────────────────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 10, s(player_name), ln=True)

    position = s(row.get("position_group_raw", "N/A"))
    club = s(current_club_display_label(row))
    league = s(current_league_label(row))
    tier_short = s(row.get("league_tier_short", get_tier_short(row.get("current_club_domestic_competition_id"))))
    contract_status = s(row.get("contract_status", "Unknown"))
    contract_badge_text = s(get_contract_badge(contract_status))

    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 7, f"{position}  |  {club}  |  {league}  |  League: {tier_short}  |  Contract: {contract_badge_text}", ln=True)
    pdf.ln(4)
    pdf.set_draw_color(0, 0, 0)
    pdf.set_line_width(0.5)
    pdf.line(18, pdf.get_y(), 192, pdf.get_y())
    pdf.ln(5)

    # ── Section 1: Key Stats ─────────────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "1. Key Stats", ln=True)
    pdf.set_font("Helvetica", "", 10)

    age_val = row.get("age_at_valuation", None)
    mins_val = row.get("minutes_last_season", None)
    curr_val = row.get("target_market_value", None)
    pred_val = row.get("predicted_value", None)
    gap_val = row.get("undervalued_pct", None)
    form_val = row.get("form_trend", "Insufficient Data")

    def _fmt_eur(v):
        try:
            return f"EUR {float(v):,.0f}"
        except (TypeError, ValueError):
            return "N/A"

    def _fmt_pct(v):
        try:
            return f"{float(v) * 100:.1f}%"
        except (TypeError, ValueError):
            return "N/A"

    stats_rows = [
        ("Age", f"{float(age_val):.1f}" if age_val is not None and not pd.isna(age_val) else "N/A"),
        ("Minutes (Last Season)", f"{float(mins_val):,.0f}" if mins_val is not None and not pd.isna(mins_val) else "N/A"),
        ("Current Market Value", _fmt_eur(curr_val)),
        ("Estimated Value", _fmt_eur(pred_val)),
        ("Value Gap", _fmt_pct(gap_val)),
        ("Form Trend", s(form_val)),
    ]
    col_w = 85
    for label, value in stats_rows:
        pdf.cell(col_w, 6, s(label) + ":", border=0)
        pdf.cell(0, 6, s(value), ln=True)
    pdf.ln(4)

    # ── Section 2: Rationale Summary ─────────────────────────────────────────
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "2. Scouting Rationale", ln=True)
    pdf.set_font("Helvetica", "", 10)
    # Strip markdown bold markers for clean PDF text
    clean_rationale = re.sub(r"\*\*(.+?)\*\*", r"\1", rationale_text)
    clean_rationale = re.sub(r"#+\s*", "", clean_rationale)
    pdf.set_x(pdf.l_margin)
    pdf.multi_cell(0, 6, s(clean_rationale))
    pdf.ln(4)

    # ── Section 3: Key Scouting Signals ──────────────────────────────────────
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "3. Key Scouting Signals", ln=True)
    pdf.set_font("Helvetica", "", 10)
    if signals_df.empty:
        pdf.cell(0, 6, "No scouting signals available.", ln=True)
    else:
        # Table header
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(55, 6, "Signal", border="B")
        pdf.cell(55, 6, "Value", border="B")
        pdf.cell(0, 6, "Scout Interpretation", border="B", ln=True)
        pdf.set_font("Helvetica", "", 9)
        for _, sig_row in signals_df.iterrows():
            signal_text = s(sig_row.get("Signal", ""))
            value_text = s(sig_row.get("Value", ""))
            interp_text = s(sig_row.get("Scout Interpretation", ""))
            # Truncate long interpretation to fit
            if len(interp_text) > 60:
                interp_text = interp_text[:57] + "..."
            pdf.cell(55, 5, signal_text[:30], border=0)
            pdf.cell(55, 5, value_text[:30], border=0)
            pdf.cell(0, 5, interp_text, ln=True)
    pdf.ln(4)

    # ── Section 4: Scout Checks Before Action ────────────────────────────────
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, "4. Scout Checks Before Action", ln=True)
    pdf.set_font("Helvetica", "", 10)
    if not scout_checks:
        pdf.cell(0, 6, "No specific checks flagged.", ln=True)
    else:
        # Explicitly reset x to left margin before each multi_cell to avoid
        # cumulative x-drift from previous cells that causes
        # "Not enough horizontal space" errors on longer lines.
        for check in scout_checks:
            bullet_line = f"- {s(check)}"
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(0, 5, bullet_line)
    pdf.ln(4)

    # ── Footer ───────────────────────────────────────────────────────────────
    pdf.set_y(-20)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_draw_color(180, 180, 180)
    pdf.line(18, pdf.get_y(), 192, pdf.get_y())
    pdf.ln(2)
    pdf.cell(
        0, 5,
        f"Generated by Football Talent Scouting Dashboard  |  {date.today().isoformat()}",
        align="C",
    )

    return bytes(pdf.output())

def calculate_scout_priority(row):
    try:
        age = row.get("age_at_valuation", 99)
        mins = row.get("minutes_last_season", 0)
        gap = row.get("undervalued_pct", 0)
        if age <= 23 and mins >= 1500 and gap >= 1.0:
            return "High Priority"
        elif age <= 25 and mins >= 900 and gap >= 0.5:
            return "Monitor"
        else:
            return "Needs Review"
    except:
        return "Needs Review"

def add_overview_context_columns(df: pd.DataFrame, club_context: pd.DataFrame) -> pd.DataFrame:
    overview_df = df.copy()
    if club_context.empty or "player_id" not in overview_df.columns:
        overview_df["Club / League"] = overview_df.apply(club_league_label, axis=1)
        return overview_df

    context_cols = [
        "player_id",
        "current_club_name",
        "current_club_domestic_competition_id",
    ]
    available_cols = [col for col in context_cols if col in club_context.columns]
    overview_df = overview_df.merge(
        club_context[available_cols].drop_duplicates(subset=["player_id"]),
        on="player_id",
        how="left",
        suffixes=("", "_context"),
    )

    for col in ["current_club_name", "current_club_domestic_competition_id"]:
        context_col = f"{col}_context"
        if context_col in overview_df.columns:
            if col in overview_df.columns:
                overview_df[col] = overview_df[col].combine_first(overview_df[context_col])
            else:
                overview_df[col] = overview_df[context_col]
            overview_df = overview_df.drop(columns=[context_col])

    overview_df["Club / League"] = overview_df.apply(club_league_label, axis=1)
    return overview_df

def build_overview_display_table(df: pd.DataFrame) -> pd.DataFrame:
    display_cols = [
        "name",
        "Club / League",
        "league_tier_short",       # Phase 8
        "contract_badge",          # Phase 9
        "form_trend",              # Phase 11
        "age_at_valuation",
        "position_group_raw",
        "primary_role",
        "minutes_last_season",
        "key_stat",                # Contextual per-position stat
        "target_market_value",
        "predicted_value",
        "undervalued_pct",
    ]
    available_cols = [col for col in display_cols if col in df.columns]

    # Fallback for Primary Role -> Position in the same column if missing
    display_df = df.copy()
    if "primary_role" not in display_df.columns and "position_group_raw" in display_df.columns:
        display_df["primary_role"] = display_df["position_group_raw"]
    elif "primary_role" in display_df.columns:
        display_df["primary_role"] = display_df["primary_role"].fillna(display_df["position_group_raw"])

    # Phase 8: backfill league_tier_short if missing
    if "league_tier_short" not in display_df.columns:
        display_df["league_tier_short"] = display_df.get(
            "current_club_domestic_competition_id", pd.Series(dtype=str)
        ).apply(get_tier_short)

    # Phase 9: backfill contract_badge if missing
    if "contract_badge" not in display_df.columns:
        display_df["contract_badge"] = display_df.get(
            "contract_status", pd.Series(dtype=str)
        ).apply(get_contract_badge)

    # Phase 11: backfill form_trend if missing
    if "form_trend" not in display_df.columns:
        display_df["form_trend"] = "Insufficient Data"

    # Map form_trend to arrow indicator for display
    form_arrow_map = {
        "Rising":            "↑ Rising",
        "Stable":            "→ Stable",
        "Declining":         "↓ Declining",
        "Insufficient Data": " -  N/A",
    }
    display_df["form_trend"] = display_df["form_trend"].map(form_arrow_map).fillna(" -  N/A")

    # Build contextual "Key Stat" per position
    def _build_key_stat(row):
        pos = row.get("position_group_raw", "")
        goals = _safe_float(row.get("goals_per_90_ls", 0))
        assists = _safe_float(row.get("assists_per_90_ls", 0))
        cards = _safe_float(row.get("cards_per_90_ls", 0))
        mins = _safe_float(row.get("minutes_last_season", 0))

        if pos == "Forward":
            if goals >= 0.3:
                return f"{goals:.2f} goals/90"
            elif assists >= 0.2:
                return f"{goals + assists:.2f} G+A/90"
            else:
                return f"{goals + assists:.2f} G+A/90"
        elif pos == "Midfielder":
            ga = goals + assists
            if assists >= 0.15:
                return f"{assists:.2f} assists/90"
            elif ga > 0:
                return f"{ga:.2f} G+A/90"
            else:
                return f"{mins:,.0f} min (creative)"
        elif pos == "Defender":
            if cards <= 0.15 and mins >= 1500:
                return f"{mins:,.0f} min · clean"
            elif cards > 0:
                return f"{mins:,.0f} min · {cards:.2f} cards/90"
            else:
                return f"{mins:,.0f} min"
        elif pos == "Goalkeeper":
            return f"{mins:,.0f} min starter"
        else:
            return f"{mins:,.0f} min"

    display_df["key_stat"] = display_df.apply(_build_key_stat, axis=1)

    display_df = display_df[available_cols].copy()

    # Calculate Priority before dropping raw valued columns
    display_df["Scout Priority"] = display_df.apply(calculate_scout_priority, axis=1)

    if "age_at_valuation" in display_df.columns:
        display_df["age_at_valuation"] = display_df["age_at_valuation"].round(1)
    if "minutes_last_season" in display_df.columns:
        display_df["minutes_last_season"] = display_df["minutes_last_season"].round(0).astype(int)
    if "target_market_value" in display_df.columns:
        display_df["target_market_value"] = pd.to_numeric(display_df["target_market_value"], errors="coerce")
    if "predicted_value" in display_df.columns:
        display_df["predicted_value"] = pd.to_numeric(display_df["predicted_value"], errors="coerce")
    if "undervalued_pct" in display_df.columns:
        display_df["undervalued_pct"] = pd.to_numeric(display_df["undervalued_pct"], errors="coerce")

    return display_df.rename(
        columns={
            "name": "Player Name",
            "league_tier_short": "League Tier",  # Phase 8
            "contract_badge": "Contract",         # Phase 9
            "form_trend": "Form",                 # Phase 11
            "age_at_valuation": "Age",
            "position_group_raw": "Position",
            "primary_role": "Primary Role",
            "minutes_last_season": "Minutes",
            "key_stat": "Key Stat",
            "target_market_value": "Current Value",
            "predicted_value": "Estimated Value",
            "undervalued_pct": "Value Gap",
        }
    )

def overview_column_config() -> dict:
    return {
        "Age": st.column_config.NumberColumn("Age", format="%.1f"),
        "Minutes": st.column_config.NumberColumn("Minutes", format="%d"),
        "Current Value": st.column_config.NumberColumn("Current Value", format="€%d"),
        "Estimated Value": st.column_config.NumberColumn("Estimated Value", format="€%d"),
        "Value Gap": st.column_config.NumberColumn("Value Gap", format="%.1f%%"),
    }

def build_similarity_display_table(res_df: pd.DataFrame, target_position: str, mode: str = "basic") -> pd.DataFrame:
    """Formats and filters columns for the similarity results table based on player position."""
    table = res_df.copy()

    def fmt_profile(value) -> str:
        if pd.isna(value):
            return np.nan
        return max(0.0, min(1.0, float(value))) * 100

    def compact_role_tags(value) -> str:
        if pd.isna(value):
            return "N/A"
        tags = [tag.strip() for tag in str(value).split(",") if tag.strip()]
        return "/".join(tags) if tags else "N/A"

    table["Age"] = table["age_at_valuation"].round(1)
    table["Age Diff"] = pd.to_numeric(table["age_difference"], errors="coerce").round(1)
    table["Height"] = pd.to_numeric(table["height_in_cm"], errors="coerce").round(0)
    table["Minutes"] = table["minutes_last_season"].round(0).astype(int)

    table["Goals/90"] = table["goals_per_90_ls"].round(2)
    table["Assists/90"] = table["assists_per_90_ls"].round(2)
    table["Cards/90"] = table["cards_per_90_ls"].round(2)

    table["Market Value"] = pd.to_numeric(table["target_market_value"], errors="coerce")
    table["Value Diff"] = pd.to_numeric(table["value_difference_vs_target"], errors="coerce")
    if "price_status" not in table.columns:
        table["price_status"] = "Unknown"
    table["Price Status"] = table["price_status"].fillna("Unknown")
    table["Predicted Value"] = pd.to_numeric(table["predicted_value"], errors="coerce")
    table["Stat Score"] = pd.to_numeric(table.get("statistical_score", table["similarity_score"]), errors="coerce") * 100
    role_score = table["role_compatibility_score"] if "role_compatibility_score" in table.columns else pd.Series(0, index=table.index)
    final_score = table["final_similarity_score"] if "final_similarity_score" in table.columns else table["similarity_score"]
    table["Role Score"] = pd.to_numeric(role_score, errors="coerce") * 100
    table["Final Score"] = pd.to_numeric(final_score, errors="coerce") * 100
    table["Primary Role"] = table["primary_role"].fillna("UNKNOWN") if "primary_role" in table.columns else "UNKNOWN"
    table["Role Tags"] = table["role_tags"].apply(compact_role_tags) if "role_tags" in table.columns else "N/A"
    table["Matched As"] = table["matched_as_role"].fillna("N/A") if "matched_as_role" in table.columns else table["Primary Role"]
    raw_fit_type = table["role_fit_type"] if "role_fit_type" in table.columns else pd.Series("N/A", index=table.index)
    table["Role Fit Type"] = [
        scout_role_fit_label(primary, tags, matched, raw_fit)
        for primary, tags, matched, raw_fit in zip(
            table["Primary Role"],
            table["Role Tags"],
            table["Matched As"],
            raw_fit_type,
        )
    ]
    role_match = table["role_matching_mode"] if "role_matching_mode" in table.columns else pd.Series("broad", index=table.index)
    table["Role Match"] = role_match.fillna("broad").str.replace("_", " ").str.title()
    table["Foot"] = table["foot"].fillna("N/A").str.title() if "foot" in table.columns else "N/A"
    table["Club / League"] = table.apply(club_league_label, axis=1)

    # Phase 8: league tier for similarity candidates
    if "league_tier_short" not in table.columns:
        table["league_tier_short"] = table.get(
            "current_club_domestic_competition_id", pd.Series(dtype=str)
        ).apply(get_tier_short)
    table["League Tier"] = table["league_tier_short"].fillna(" - ")

    # Phase 9: contract status for similarity candidates
    if "contract_badge" not in table.columns:
        table["contract_badge"] = table.get(
            "contract_status", pd.Series(dtype=str)
        ).apply(get_contract_badge)
    table["Contract"] = table["contract_badge"].fillna("❓ Unknown")

    table["Cheaper?"] = table["is_cheaper"].apply(lambda x: "✅ Yes" if x else "❌ No")
    table["Younger?"] = table["is_younger"].apply(lambda x: "✅ Yes" if x else "❌ No")
    table["Undervalued?"] = table["is_undervalued"].apply(lambda x: "💎 Yes" if x else " -  No")

    base_cols = [
        "name", "Club / League", "League Tier", "Contract", "Primary Role", "Role Tags",
        "Matched As", "Role Fit Type", "Foot", "Age", "Age Diff",
    ]
    value_cols = [
        "Market Value", "Value Diff", "Price Status", "Predicted Value", "Stat Score",
        "Role Score", "Final Score", "Cheaper?", "Younger?", "Undervalued?"
    ]

    if mode == "enriched":
        if target_position == "Forward":
            if "profile_goal_threat" in table.columns:
                table["Goal Threat"] = table["profile_goal_threat"].apply(fmt_profile)
                table["Chance Creation"] = table["profile_chance_creation"].apply(fmt_profile)
                table["1v1 Threat"] = table["profile_1v1_threat"].apply(fmt_profile)
                table["Box Involvement"] = table["profile_box_involvement"].apply(fmt_profile)
                table["Ball Progression"] = table["profile_ball_progression"].apply(fmt_profile)
                adv_cols = ["Goal Threat", "Chance Creation", "1v1 Threat", "Box Involvement", "Ball Progression"]
            else:
                table["xG"] = table["adv_Expected_xG"].round(2)
                table["SoT%"] = table["adv_Standard_SoT%"].round(1)
                table["SCA90"] = table["adv_SCA_SCA90"].round(2)
                table["Take-On%"] = table["adv_Take-Ons_Succ%"].round(1)
                adv_cols = ["xG", "SoT%", "SCA90", "Take-On%"]
        elif target_position == "Midfielder":
            if "profile_ball_progression" in table.columns:
                table["Ball Progression"] = table["profile_ball_progression"].apply(fmt_profile)
                table["Chance Creation"] = table["profile_chance_creation"].apply(fmt_profile)
                table["Possession Quality"] = table["profile_possession_quality"].apply(fmt_profile)
                table["Defensive Activity"] = table["profile_defensive_activity"].apply(fmt_profile)
                table["Carrying"] = table["profile_carrying"].apply(fmt_profile)
                adv_cols = ["Ball Progression", "Chance Creation", "Possession Quality", "Defensive Activity", "Carrying"]
            else:
                table["Key Passes"] = table["adv_KP_"].round(2)
                table["Progressive Passes"] = table["adv_Progression_PrgP"].round(2)
                table["Tackles Won"] = table["adv_Tackles_TklW"].round(2)
                table["Interceptions"] = table["adv_Int_"].round(2)
                adv_cols = ["Key Passes", "Progressive Passes", "Tackles Won", "Interceptions"]
        elif target_position == "Defender":
            if "profile_defending_volume" in table.columns:
                table["Defending Volume"] = table["profile_defending_volume"].apply(fmt_profile)
                table["Duel Strength"] = table["profile_duel_strength"].apply(fmt_profile)
                table["Aerial Strength"] = table["profile_aerial_strength"].apply(fmt_profile)
                table["Build-up Support"] = table["profile_build_up_support"].apply(fmt_profile)
                table["Risk Control"] = table["profile_risk_control"].apply(fmt_profile)
                adv_cols = ["Defending Volume", "Duel Strength", "Aerial Strength", "Build-up Support", "Risk Control"]
            else:
                table["Blocks"] = table["adv_Blocks_Blocks"].round(2)
                table["Clearances"] = table["adv_Clr_"].round(2)
                table["Aerial Win%"] = table["adv_Aerial Duels_Won%"].round(1)
                table["Tackle Win%"] = table["adv_Challenges_Tkl%"].round(1)
                adv_cols = ["Blocks", "Clearances", "Aerial Win%", "Tackle Win%"]
        else: # Goalkeeper
            if "profile_distribution" in table.columns:
                table["Distribution"] = table["profile_distribution"].apply(fmt_profile)
                table["Recoveries"] = table["profile_recoveries"].apply(fmt_profile)
                table["Aerial Control"] = table["profile_aerial_control"].apply(fmt_profile)
                adv_cols = ["Distribution", "Recoveries", "Aerial Control"]
            else:
                table["Pass Cmp%"] = table["adv_Total_Cmp%"].round(1)
                table["Recoveries"] = table["adv_Performance_Recov"].round(2)
                table["Aerial Win%"] = table["adv_Aerial Duels_Won%"].round(1)
                adv_cols = ["Pass Cmp%", "Recoveries", "Aerial Win%"]

        selected_cols = base_cols + adv_cols + value_cols
    else:
        # Basic mode
        if target_position == "Goalkeeper":
            selected_cols = base_cols + ["Height", "Minutes"] + value_cols
        elif target_position == "Defender":
            selected_cols = base_cols + ["Height", "Minutes", "Cards/90"] + value_cols
        else: # Forward / Midfielder
            selected_cols = base_cols + ["Goals/90", "Assists/90", "Cards/90"] + value_cols

    selected_cols = [col for col in selected_cols if col in table.columns]
    final_table = table[selected_cols].copy()
    final_table = final_table.rename(
        columns={
            "name": "Player Name",
            "Role Match": "Match Type",
            "Market Value": "Current Value",
            "Predicted Value": "Estimated Value",
            "Stat Score": "Playing Profile Match",
            "Role Score": "Role Fit",
            "Final Score": "Overall Fit",
            "Undervalued?": "Scout Flag",
        }
    )
    return final_table

def similarity_column_config() -> dict:
    percent_cols = [
        "Playing Profile Match",
        "Role Fit",
        "Overall Fit",
        "SoT%",
        "Take-On%",
        "Aerial Win%",
        "Tackle Win%",
    ]
    profile_cols = [
        "Goal Threat",
        "Chance Creation",
        "1v1 Threat",
        "Box Involvement",
        "Ball Progression",
        "Possession Quality",
        "Defensive Activity",
        "Carrying",
        "Defending Volume",
        "Duel Strength",
        "Aerial Strength",
        "Build-up Support",
        "Risk Control",
        "Distribution",
        "Recoveries",
        "Aerial Control",
    ]
    config = {
        "Age": st.column_config.NumberColumn("Age", format="%.1f"),
        "Age Diff": st.column_config.NumberColumn("Age Diff", format="%+.1f yrs"),
        "Height": st.column_config.NumberColumn("Height", format="%.0f cm"),
        "Minutes": st.column_config.NumberColumn("Minutes", format="%d"),
        "Current Value": st.column_config.NumberColumn("Current Value", format="€%d"),
        "Value Diff": st.column_config.NumberColumn("Value Diff", format="€%+d"),
        "Estimated Value": st.column_config.NumberColumn("Estimated Value", format="€%d"),
        "Goals/90": st.column_config.NumberColumn("Goals/90", format="%.2f"),
        "Assists/90": st.column_config.NumberColumn("Assists/90", format="%.2f"),
        "Cards/90": st.column_config.NumberColumn("Cards/90", format="%.2f"),
        "xG": st.column_config.NumberColumn("xG", format="%.2f"),
        "SCA90": st.column_config.NumberColumn("SCA90", format="%.2f"),
        "Key Passes": st.column_config.NumberColumn("Key Passes", format="%.2f"),
        "Progressive Passes": st.column_config.NumberColumn("Progressive Passes", format="%.2f"),
        "Tackles Won": st.column_config.NumberColumn("Tackles Won", format="%.2f"),
        "Interceptions": st.column_config.NumberColumn("Interceptions", format="%.2f"),
        "Blocks": st.column_config.NumberColumn("Blocks", format="%.2f"),
        "Clearances": st.column_config.NumberColumn("Clearances", format="%.2f"),
        "Pass Cmp%": st.column_config.NumberColumn("Pass Cmp%", format="%.1f%%"),
    }
    for col in percent_cols:
        config[col] = st.column_config.NumberColumn(col, format="%.2f%%")
    for col in profile_cols:
        config[col] = st.column_config.ProgressColumn(
            col,
            min_value=0,
            max_value=100,
            format="%.0f/100",
        )
    return config

def format_percentage_points(value) -> str:
    if pd.isna(value):
        return "N/A"
    return f"{float(value) * 100:+.1f} pp"

def backtest_column_config() -> dict:
    return {
        "Snapshot": st.column_config.DateColumn("Snapshot"),
        "Age": st.column_config.NumberColumn("Age", format="%.1f"),
        "Minutes": st.column_config.NumberColumn("Minutes", format="%d"),
        "Current Value": st.column_config.NumberColumn("Current Value", format="€%d"),
        "Estimated Value": st.column_config.NumberColumn("Estimated Value", format="€%d"),
        "Value Gap": st.column_config.NumberColumn("Value Gap", format="%.1f%%"),
        "Future 12M Value": st.column_config.NumberColumn("Future 12M Value", format="€%d"),
        "12M Growth": st.column_config.NumberColumn("12M Growth", format="%+.1f%%"),
    }

def build_backtest_display_table(candidates_df: pd.DataFrame) -> pd.DataFrame:
    if candidates_df.empty:
        return pd.DataFrame()

    table = candidates_df.copy()
    display_cols = [
        "snapshot_date",
        "name",
        "position_group_raw",
        "primary_role",
        "age_at_snapshot",
        "minutes_last_365",
        "current_market_value",
        "predicted_value",
        "value_gap",
        "future_value_12m",
        "growth_12m_pct",
        "hit_12m",
    ]
    available_cols = [col for col in display_cols if col in table.columns]
    table = table[available_cols].copy()
    if "snapshot_date" in table.columns:
        table["snapshot_date"] = pd.to_datetime(table["snapshot_date"], errors="coerce")
    for col in [
        "age_at_snapshot",
        "minutes_last_365",
        "current_market_value",
        "predicted_value",
        "value_gap",
        "future_value_12m",
        "growth_12m_pct",
    ]:
        if col in table.columns:
            table[col] = pd.to_numeric(table[col], errors="coerce")
    if "value_gap" in table.columns:
        table["value_gap"] = table["value_gap"] * 100
    if "growth_12m_pct" in table.columns:
        table["growth_12m_pct"] = table["growth_12m_pct"] * 100
    if "hit_12m" in table.columns:
        table["hit_12m"] = table["hit_12m"].apply(lambda value: "Yes" if scalar_bool(value) else "No")

    return table.rename(
        columns={
            "snapshot_date": "Snapshot",
            "name": "Player",
            "position_group_raw": "Position",
            "primary_role": "Primary Role",
            "age_at_snapshot": "Age",
            "minutes_last_365": "Minutes",
            "current_market_value": "Current Value",
            "predicted_value": "Estimated Value",
            "value_gap": "Value Gap",
            "future_value_12m": "Future 12M Value",
            "growth_12m_pct": "12M Growth",
            "hit_12m": "12M Hit",
        }
    )


# =========================
# Resource Loading
# =========================
@st.cache_data
def load_scouting_data() -> pd.DataFrame:
    """Load scouting output for the overview tab."""
    candidate_paths = [
        os.path.join("outputs", "undervalued_candidates_overall.csv"),
        os.path.join("outputs", "hidden_gems_overall.csv"),
    ]
    for file_path in candidate_paths:
        if os.path.exists(file_path):
            df = pd.read_csv(file_path)
            # Phase 8: ensure league tier columns are present
            # (backfill for files generated before Phase 8 pipeline run)
            if "league_tier" not in df.columns:
                df = enrich_league_tier(df, competition_col="current_club_domestic_competition_id")
            # Phase 9: ensure contract columns are present
            if "contract_status" not in df.columns:
                df = enrich_contract(df)
            # Phase 11: ensure form_trend column is present
            if "form_trend" not in df.columns:
                df["form_trend"] = "Insufficient Data"
            return df
    return pd.DataFrame()

@st.cache_data
def load_all_predictions() -> pd.DataFrame:
    """Load the full modeled player pool for rationale search."""
    pred_path = os.path.join("outputs", "predictions_per_player.csv")
    if os.path.exists(pred_path):
        df = pd.read_csv(pred_path)
        # Phase 8: ensure league tier columns are present
        if "league_tier" not in df.columns:
            df = enrich_league_tier(df, competition_col="current_club_domestic_competition_id")
        # Phase 9: ensure contract columns are present
        if "contract_status" not in df.columns:
            df = enrich_contract(df)
        # Phase 11: ensure form_trend column is present
        if "form_trend" not in df.columns:
            df["form_trend"] = "Insufficient Data"
        return df
    return pd.DataFrame()

@st.cache_data
def load_temporal_backtest_outputs() -> dict:
    """Load Phase 7 temporal validation outputs for the dashboard."""
    base_path = os.path.join("outputs", "temporal_validation")
    file_map = {
        "summary": "temporal_backtest_summary.csv",
        "candidates": "temporal_backtest_candidates.csv",
        "by_value_gap": "success_rate_by_value_gap_bucket.csv",
        "by_position": "success_rate_by_position.csv",
        "by_age_group": "success_rate_by_age_group.csv",
    }
    outputs = {}
    missing = []
    for key, filename in file_map.items():
        path = os.path.join(base_path, filename)
        if os.path.exists(path):
            outputs[key] = pd.read_csv(path)
        else:
            outputs[key] = pd.DataFrame()
            missing.append(path)

    report_path = os.path.join(base_path, "temporal_validation_report.json")
    if os.path.exists(report_path):
        with open(report_path, "r", encoding="utf-8") as f:
            outputs["report"] = json.load(f)
    else:
        outputs["report"] = {}
        missing.append(report_path)

    outputs["missing"] = missing
    return outputs

@st.cache_data
def load_current_club_context() -> pd.DataFrame:
    """Load pre-computed club context for display-only dashboard enrichment.

    Uses the lightweight pre-computed file (data/processed/club_context.csv)
    instead of loading large raw files (appearances.csv, player_valuations.csv)
    at runtime. This reduces dashboard startup from ~170 MB to <1 MB of I/O.
    """
    # Prefer pre-computed context (generated by precompute_context.py or pipeline)
    precomputed_path = os.path.join("data", "processed", "club_context.csv")
    if os.path.exists(precomputed_path):
        return pd.read_csv(precomputed_path)

    # Fallback: extract from featured_players if pre-computed file is missing
    context_path = os.path.join("data", "processed", "featured_players.csv")
    if not os.path.exists(context_path):
        return pd.DataFrame()

    context_cols = [
        "player_id",
        "current_club_name",
        "current_club_domestic_competition_id",
    ]
    df_context = pd.read_csv(context_path)
    available = [col for col in context_cols if col in df_context.columns]
    return df_context[available].drop_duplicates(subset=["player_id"]).copy()

def load_similarity_engine():
    """Load the similarity search model."""
    engine_path = os.path.join("outputs", "models", "similarity_engine.pkl")
    if os.path.exists(engine_path):
        return joblib.load(engine_path)
    return None


# =========================
# Header
# =========================
st.markdown(
    """
    <div class="app-hero">
        <div class="app-title">⚽ Football Talent Scouting Dashboard</div>
        <div class="app-subtitle">
            Find undervalued players, understand the scouting rationale, and compare role-aware alternatives.
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

tab_scouting, tab_rationale, tab_similarity, tab_backtest = st.tabs([
    "💎 Find Undervalued Players",
    "🧠 Why This Player?",
    "🔍 Player Alternatives",
    "📈 Backtest Results",
])

# =========================
# TAB 1: CANDIDATE SHORTLIST
# =========================
with tab_scouting:
    df = load_scouting_data()
    club_context_df = load_current_club_context()

    if df.empty:
        st.error("No scouting data found. Please run the pipeline first.")
    else:
        st.markdown("#### Candidate Filters")
        st.caption("Use these filters to narrow the undervalued shortlist before reviewing candidates.")

        positions = sorted(df["position_group_raw"].dropna().unique().tolist())
        min_age, max_age = int(df["age_at_valuation"].min()), int(df["age_at_valuation"].max())
        min_val, max_val = int(df["target_market_value"].min()), int(df["target_market_value"].max())
        min_mins, max_mins = int(df["minutes_last_season"].min()), int(df["minutes_last_season"].max())

        filter_col_1, filter_col_2 = st.columns([1.2, 1])
        with filter_col_1:
            selected_positions = st.multiselect("Position Group", positions, default=positions)
        with filter_col_2:
            selected_age = st.slider("Age Range", min_age, max_age, (min_age, 25))

        filter_col_3, filter_col_4 = st.columns(2)
        with filter_col_3:
            selected_val = st.slider(
                "Max Market Value (€)",
                min_val,
                max_val,
                min(max(5000000, min_val), max_val),
                step=500000,
            )
        with filter_col_4:
            selected_mins = st.slider(
                "Minimum Minutes Played",
                min_mins,
                max_mins,
                min(max(900, min_mins), max_mins),
            )

        # Phase 8: League Tier filter
        filter_col_5, filter_col_6 = st.columns([1.2, 1])
        with filter_col_5:
            tier_options = {
                "All Tiers": [1, 2, 3, 4],
                "Tier 1  -  Elite only": [1],
                "Tier 1 + 2 (Elite & Strong)": [1, 2],
                "Tier 3 + 4 (Competitive & Developing)": [3, 4],
            }
            selected_tier_label = st.selectbox(
                "League Tier",
                options=list(tier_options.keys()),
                index=0,
                help="Filter candidates by the competitive level of their current league.",
            )
            selected_tiers = tier_options[selected_tier_label]

        # Phase 9: Contract Status filter
        with filter_col_6:
            contract_filter_options = ["All Contracts"] + list(CONTRACT_STATUS_LABELS.values())
            selected_contract_label = st.selectbox(
                "Contract Status",
                options=contract_filter_options,
                index=0,
                help="Filter by contract situation. 'Expiring' players may be available at lower cost.",
            )
            # Map display label back to status key
            label_to_key = {v: k for k, v in CONTRACT_STATUS_LABELS.items()}
            selected_contract_key = label_to_key.get(selected_contract_label, None)

        # Apply Filters
        contract_mask = (
            df["contract_status"] == selected_contract_key
            if selected_contract_key and "contract_status" in df.columns
            else pd.Series(True, index=df.index)
        )
        filtered_df = df[
            (df["position_group_raw"].isin(selected_positions)) &
            (df["age_at_valuation"].between(selected_age[0], selected_age[1])) &
            (df["target_market_value"] <= selected_val) &
            (df["minutes_last_season"] >= selected_mins) &
            (df["league_tier"].isin(selected_tiers) if "league_tier" in df.columns else True) &
            contract_mask
        ].copy()

        filtered_df = filtered_df.sort_values(by="undervalued_pct", ascending=False)
        st.session_state["active_candidate_shortlist"] = filtered_df.copy()

        # Metrics
        st.markdown("### 📊 Shortlist Summary")
        k1, k2, k3, k4 = st.columns(4)
        with k1: kpi_card("Shortlisted Players", f"{len(filtered_df):,}")
        with k2: kpi_card("Average Age", f"{filtered_df['age_at_valuation'].mean():.1f} yrs" if not filtered_df.empty else "N/A")
        with k3: kpi_card("Average Value Gap", f"{filtered_df['undervalued_pct'].mean()*100:.1f}%" if not filtered_df.empty else "N/A")
        with k4: kpi_card("Median Current Value", f"€{filtered_df['target_market_value'].median():,.0f}" if not filtered_df.empty else "N/A")

        methodology_box("This shortlist highlights players whose recent output appears stronger than their current market value. Use it as a first-pass scouting filter, not a final recommendation.")
        st.divider()

        if not filtered_df.empty:
            overview_df = add_overview_context_columns(filtered_df, club_context_df)
            st.markdown("### 📈 Candidate Map")
            c1, c2 = st.columns(2)

            with c1:
                fig1 = px.scatter(
                    filtered_df, x="age_at_valuation", y="undervalued_pct",
                    color="position_group_raw", size="target_market_value", hover_name="name",
                    title="Age vs Value Gap", template=PLOTLY_TEMPLATE,
                    labels={
                        "age_at_valuation": "Age",
                        "undervalued_pct": "Value Gap",
                        "position_group_raw": "Position",
                        "target_market_value": "Current Value (€)",
                        "name": "Player Name",
                    },
                )
                fig1.update_yaxes(tickformat=".0%")
                st.plotly_chart(fig1, width="stretch")

            with c2:
                fig2 = px.scatter(
                    filtered_df, x="target_market_value", y="predicted_value",
                    color="position_group_raw", hover_name="name",
                    title="Current vs Estimated Value", template=PLOTLY_TEMPLATE,
                    labels={
                        "target_market_value": "Current Value (€)",
                        "predicted_value": "Estimated Value (€)",
                        "position_group_raw": "Position",
                        "name": "Player Name",
                    },
                )
                max_l = max(filtered_df["target_market_value"].max(), filtered_df["predicted_value"].max())
                fig2.add_shape(type="line", line=dict(dash="dash", color=REFERENCE_LINE_COLOR), x0=0, y0=0, x1=max_l, y1=max_l)
                st.plotly_chart(fig2, width="stretch")
                st.caption("Points above the dashed line indicate players estimated to be worth more than their current market value.")

            st.divider()
            st.markdown("### 📋 Top Scouting Leads")

            display_df = build_overview_display_table(overview_df)

            st.dataframe(
                display_df,
                width="stretch",
                hide_index=True,
                column_config=overview_column_config(),
            )
            download_df = add_download_context_columns(overview_df)
            # Phase 12: active filter summary for metadata sheet
            _active_filters = {
                "Positions": ", ".join(selected_positions) if selected_positions else "All",
                "Age": f"{selected_age[0]}-{selected_age[1]}",
                "Max Market Value": f"EUR {selected_val:,}",
                "Min Minutes": str(selected_mins),
                "League Tier": selected_tier_label,
                "Contract Status": selected_contract_label,
            }
            if _OPENPYXL_AVAILABLE:
                _excel_bytes = build_shortlist_excel(overview_df, active_filters=_active_filters)
                st.download_button(
                    label="💾 Download Filtered Shortlist (Excel)",
                    data=_excel_bytes,
                    file_name="filtered_undervalued_shortlist.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.warning("openpyxl is not installed  -  Excel export unavailable. Install it with `pip install openpyxl`.")
                st.download_button(
                    label="💾 Download Filtered Shortlist (CSV fallback)",
                    data=download_df.to_csv(index=False),
                    file_name="filtered_undervalued_shortlist.csv",
                    mime="text/csv",
                )

            # ── Phase 12: Session Watchlist ──────────────────────────────────
            if "watchlist" not in st.session_state:
                st.session_state["watchlist"] = []

            st.divider()
            st.markdown("#### ➕ Add to Watchlist")
            _shortlist_names = overview_df["name"].dropna().tolist() if "name" in overview_df.columns else []
            _selected_for_watchlist = st.multiselect(
                "Select players to add to watchlist",
                options=_shortlist_names,
                default=[],
                key="watchlist_multiselect",
                help="Players are saved for the current session only.",
            )
            if st.button("Add Selected to Watchlist", key="btn_add_watchlist"):
                _added = 0
                for _pname in _selected_for_watchlist:
                    if _pname not in st.session_state["watchlist"]:
                        st.session_state["watchlist"].append(_pname)
                        _added += 1
                if _added:
                    st.success(f"Added {_added} player(s) to watchlist.")
                else:
                    st.info("Selected players are already in the watchlist.")

            with st.expander(f"📌 My Watchlist ({len(st.session_state['watchlist'])} players)", expanded=False):
                if not st.session_state["watchlist"]:
                    st.info("Your watchlist is empty. Use the multiselect above to add players.")
                else:
                    _wl_names = st.session_state["watchlist"]
                    _wl_df = df[df["name"].isin(_wl_names)].copy() if "name" in df.columns else pd.DataFrame()
                    if not _wl_df.empty:
                        _wl_display = add_overview_context_columns(_wl_df, club_context_df)
                        _wl_table = build_overview_display_table(_wl_display)
                        st.dataframe(_wl_table, width="stretch", hide_index=True, column_config=overview_column_config())
                    else:
                        st.write("Watchlisted players: " + ", ".join(_wl_names))

                    _col_clear, _col_export = st.columns(2)
                    with _col_clear:
                        if st.button("🗑️ Clear Watchlist", key="btn_clear_watchlist"):
                            st.session_state["watchlist"] = []
                            st.rerun()
                    with _col_export:
                        if not _wl_df.empty:
                            if _OPENPYXL_AVAILABLE:
                                _wl_excel = build_shortlist_excel(_wl_display, active_filters={"Source": "Watchlist"})
                                st.download_button(
                                    label="💾 Export Watchlist (Excel)",
                                    data=_wl_excel,
                                    file_name="watchlist_export.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="btn_export_watchlist",
                                )
                            else:
                                st.download_button(
                                    label="💾 Export Watchlist (CSV)",
                                    data=add_download_context_columns(_wl_display).to_csv(index=False),
                                    file_name="watchlist_export.csv",
                                    mime="text/csv",
                                    key="btn_export_watchlist_csv",
                                )
            disclaimer_box(
                "Statistical leads should be validated by professional human scouting. "
                "The 'Key Stat' column shows the most relevant available metric per position "
                "(goals+assists for attackers, minutes and discipline for defenders, minutes for goalkeepers). "
                "Advanced metrics such as xG, progressive passes, tackles, and save percentage are only available for a small subset "
                "of Top 5 League players  -  use the Player Alternatives tab for detailed profile comparisons."
            )
        else:
            st.warning("Adjust filters to see candidates.")


# =========================
# TAB 2: PLAYER ALTERNATIVES
# =========================
with tab_similarity:
    engine = load_similarity_engine()

    if engine is None:
        st.error("Similarity engine not found. Please run `python src/similarity.py` first.")
    else:
        st.markdown("### 🔍 Player Alternatives")
        st.markdown("Start from the closest role-aware playing-profile matches, then apply recruitment filters to fit your scouting brief.")
        # Player Selection
        all_players = engine.data.sort_values(by="name")
        player_options = {row["name"]: row["player_id"] for _, row in all_players.iterrows()}

        target_player_name = st.selectbox(
            "Select target player",
            options=list(player_options.keys()),
            index=list(player_options.keys()).index("Aaron Ramsdale") if "Aaron Ramsdale" in player_options else 0
        )
        target_player_id = player_options[target_player_name]

        target_data = engine.data[engine.data['player_id'] == target_player_id].iloc[0]

        # Defensive check for enrichment metadata migration.
        if 'enriched_available' in target_data:
            has_enriched = scalar_bool(target_data['enriched_available'])
        elif 'match_status' in target_data:
            has_enriched = target_data['match_status'] == 'matched'
        else:
            # Fallback if engine artifact is outdated
            has_enriched = False
            st.error("⚠️ Similarity data is outdated. Please re-run `python src/similarity.py` to enable enriched mode.")
        has_role_metadata = scalar_bool(target_data.get('role_metadata_available', False))

        # Profile definition settings
        col_s1, col_s2 = st.columns([1, 1])
        with col_s1:
            role_mode_label = st.selectbox(
                "Role Matching Mode",
                ["Similar Roles", "Same Role Only", "Same Position Group"],
                index=0,
                help="Similar Roles uses the role compatibility matrix. Same Role Only requires the same explicit role tag. Same Position Group falls back to position-group behavior."
            )
            role_mode_map = {
                "Similar Roles": "compatible",
                "Same Role Only": "exact",
                "Same Position Group": "broad",
            }
            role_mode_value = role_mode_map[role_mode_label]
        with col_s2:
            if has_enriched:
                sim_engine_mode = st.selectbox(
                    "Profile Mode",
                    ["Advanced Profile", "Standard Profile"],
                    index=0,
                    help="Advanced Profile uses additional league performance data where safely matched. Standard Profile uses only basic stats."
                )
                actual_mode = "enriched" if "Advanced" in sim_engine_mode else "basic"
                st.markdown('<span class="advanced-stats-badge">Advanced Stats Available</span>', unsafe_allow_html=True)
            else:
                st.selectbox("Profile Mode", ["Standard Profile"], disabled=True)
                actual_mode = "basic"
                st.markdown('<span class="basic-stats-badge">Core Stats Only</span>', unsafe_allow_html=True)

        # Show target profile
        st.markdown("#### 👤 Target Player")
        target_role = target_data.get('primary_role', 'UNKNOWN') if has_role_metadata else "UNKNOWN"
        target_tags = format_role_tags(target_data.get('role_tags', None))
        target_foot = format_foot_profile(target_data.get('foot', None), has_role_metadata)
        role_chips = [f"Playing Role: {target_role}", target_foot]
        if not role_tags_are_redundant(str(target_role), target_tags):
            role_chips.insert(1, f"Tags: {target_tags}")
        role_metadata_label = "Metadata available" if has_role_metadata else "Broad-position fallback"
        role_chips.append(role_metadata_label)
        match_method = target_data.get('match_method', 'N/A' if has_enriched else 'none')
        match_score = target_data.get('match_score', 0)
        match_method_label = format_match_method(match_method)
        match_confidence_label = format_match_confidence(match_score) if has_enriched else "N/A"
        stats_season_label = format_stats_season(target_data.get('kaggle_season', None))
        similarity_mode_label = "Advanced Profile" if actual_mode == "enriched" else "Standard Profile"

        if has_enriched and actual_mode == "enriched":
            data_context = f"Advanced player data available. Stats season: {stats_season_label}. Player data matched with high confidence."
        elif has_enriched:
            data_context = f"Advanced player data available. Stats season: {stats_season_label}. Player data matched with high confidence."
        else:
            data_context = "No advanced data match. Using standard stats."

        # Phase 8: build league tier badge for target player
        target_comp_id = target_data.get("current_club_domestic_competition_id", None)
        target_tier_badge = league_tier_badge(target_comp_id)
        target_tier_note = get_scout_note(target_comp_id)

        # Phase 9: build contract badge for target player
        target_contract_status = target_data.get("contract_status", "Unknown")
        target_contract_badge_html = contract_status_badge(target_contract_status)
        target_contract_note = get_contract_scout_note(target_contract_status)

        st.markdown(
            f"""
            <div class="target-panel">
                <div class="target-name">{html.escape(str(target_player_name))} {target_tier_badge} {target_contract_badge_html}</div>
                <div class="target-summary">
                    {html.escape(str(target_data['position_group_raw']))} · Playing Role: {html.escape(str(target_role))} · {html.escape(target_foot)}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if target_tier_note:
            st.caption(f"🌍 **League Context:** {target_tier_note}")
        if target_contract_note:
            st.caption(f"📋 **Contract:** {target_contract_note}")

        # Compact info line replaces the 4 Position/Age/Value/Minutes cards
        render_info_line([
            ("Position", target_data.get("position_group_raw", "N/A")),
            ("Age", f"{target_data['age_at_valuation']:.1f}"),
            ("Current Value", format_currency(target_data.get('target_market_value', 0))),
            ("Minutes Played", f"{_safe_float(target_data.get('minutes_last_season', 0)):,.0f}"),
        ])

        st.markdown("##### Role Profile")
        st.markdown(render_chip_row(role_chips), unsafe_allow_html=True)

        st.markdown("##### Data Context")
        st.caption(f"**Data Used:** {data_context}")
        if has_enriched:
            with st.expander("View data matching details", expanded=False):
                st.markdown(f"Match Method: `{match_method_label}` | Match Confidence: `{match_confidence_label}`")
                details_df = pd.DataFrame(
                    [
                        {"Field": "Profile Mode", "Value": similarity_mode_label},
                        {"Field": "Match Method", "Value": match_method_label},
                        {"Field": "Match Confidence", "Value": match_confidence_label},
                        {"Field": "Stats Season", "Value": stats_season_label},
                    ]
                )
                st.dataframe(details_df, width="stretch", hide_index=True)
                st.caption("Match confidence refers to the player metadata matching process between datasets. It is not the similarity score.")
                fuzzy_note = fuzzy_confidence_note(match_method, match_score)
                if fuzzy_note:
                    st.caption(fuzzy_note)


        st.divider()

        # ── Target Player Statistics (for comparison with alternatives) ─────
        # Show the target player's core stats + profile metrics in the same
        # format as the alternatives table so scouts can compare directly.
        target_position = target_data["position_group_raw"]
        target_stats_rows = []

        # Core stats (always available)
        target_stats_rows.extend([
            ("Goals/90", f"{_safe_float(target_data.get('goals_per_90_ls', 0)):.2f}"),
            ("Assists/90", f"{_safe_float(target_data.get('assists_per_90_ls', 0)):.2f}"),
            ("Cards/90", f"{_safe_float(target_data.get('cards_per_90_ls', 0)):.2f}"),
        ])

        # Profile metrics (enriched mode)  -  position-specific
        if actual_mode == "enriched":
            profile_map = {
                "Forward": [
                    ("Goal Threat", "profile_goal_threat"),
                    ("Chance Creation", "profile_chance_creation"),
                    ("1v1 Threat", "profile_1v1_threat"),
                    ("Box Involvement", "profile_box_involvement"),
                    ("Ball Progression", "profile_ball_progression"),
                ],
                "Midfielder": [
                    ("Ball Progression", "profile_ball_progression"),
                    ("Chance Creation", "profile_chance_creation"),
                    ("Possession Quality", "profile_possession_quality"),
                    ("Defensive Activity", "profile_defensive_activity"),
                    ("Carrying", "profile_carrying"),
                ],
                "Defender": [
                    ("Defending Volume", "profile_defending_volume"),
                    ("Duel Strength", "profile_duel_strength"),
                    ("Aerial Strength", "profile_aerial_strength"),
                    ("Build-up Support", "profile_build_up_support"),
                    ("Risk Control", "profile_risk_control"),
                ],
                "Goalkeeper": [
                    ("Distribution", "profile_distribution"),
                    ("Recoveries", "profile_recoveries"),
                    ("Aerial Control", "profile_aerial_control"),
                ],
            }
            for label, col in profile_map.get(target_position, []):
                val = target_data.get(col, None)
                if val is None or pd.isna(val):
                    target_stats_rows.append((label, "N/A"))
                else:
                    target_stats_rows.append((label, f"{float(val) * 100:.0f}/100"))

        st.markdown(f"##### 📊 Target Player Statistics · {target_player_name}")
        st.caption("Use these numbers as a reference when comparing the alternatives table below.")

        # Render as a horizontal table (single row)
        target_stats_df = pd.DataFrame([{label: value for label, value in target_stats_rows}])
        st.dataframe(target_stats_df, width="stretch", hide_index=True)

        st.divider()

        # Similarity Search
        st.caption(f"Alternatives for **{target_player_name}** using **{actual_mode.upper()}** statistical mode and **{role_mode_label}** role logic.")
        if role_mode_value == "broad":
            st.warning("⚠️ Broad Position Group uses fallback position logic and may produce less role-specific matches.")
        if not has_role_metadata:
            st.info("Role metadata is not available for this target, so the engine falls back to broad position-group matching.")

        # Get a larger pool first so recruitment filters do not hide good matches too early.
        similar_pool_df = engine.find_similar(
            target_player_id,
            top_n=300,
            same_position=True,
            mode=actual_mode,
            role_mode=role_mode_value,
            role_threshold=COMPATIBLE_ROLE_THRESHOLD,
        )

        if similar_pool_df is not None and not similar_pool_df.empty:
            similar_pool_df = similar_pool_df.copy()
            target_value_raw = target_data.get("target_market_value", 0)
            target_value = 0.0 if pd.isna(target_value_raw) else float(target_value_raw)
            similar_pool_df["price_status"] = similar_pool_df["value_difference_vs_target"].apply(
                lambda diff: price_status_label(diff, target_value)
            )

            score_col = "final_similarity_score" if "final_similarity_score" in similar_pool_df.columns else "similarity_score"
            role_score_col = "role_compatibility_score" if "role_compatibility_score" in similar_pool_df.columns else None
            similar_pool_df[score_col] = pd.to_numeric(similar_pool_df[score_col], errors="coerce").fillna(0)
            if role_score_col:
                similar_pool_df[role_score_col] = pd.to_numeric(similar_pool_df[role_score_col], errors="coerce").fillna(0)

            st.markdown("#### Recruitment Filters")
            st.caption("These filters do not change the similarity score. They narrow the ranked alternatives after the model finds similar players.")

            age_series = pd.to_numeric(similar_pool_df["age_at_valuation"], errors="coerce")
            max_age_limit = int(max(16, np.ceil(age_series.max() if age_series.notna().any() else 35)))
            default_max_age = min(35, max_age_limit)

            value_series = pd.to_numeric(similar_pool_df["target_market_value"], errors="coerce")
            max_pool_value = float(value_series.max()) if value_series.notna().any() else max(target_value, 500_000)
            max_value_limit = int(np.ceil(max(max_pool_value, target_value, 500_000) / 500_000) * 500_000)
            default_max_value = max_value_limit

            f1, f2, f3 = st.columns(3)
            with f1:
                max_age = st.slider("Max Age", min_value=16, max_value=max_age_limit, value=default_max_age, step=1)
            with f2:
                max_value = st.slider(
                    "Max Current Value (€)",
                    min_value=0,
                    max_value=max_value_limit,
                    value=default_max_value,
                    step=500_000,
                )
            with f3:
                max_results = st.selectbox("Results", [10, 25, 50], index=0)

            f5, f6, f7, f8, f9 = st.columns(5)
            with f5:
                min_overall_fit = st.slider("Min Overall Fit", min_value=0, max_value=100, value=90, step=1)
            with f6:
                min_role_fit = st.slider("Min Role Fit", min_value=0, max_value=100, value=int(COMPATIBLE_ROLE_THRESHOLD * 100), step=1)
            with f7:
                younger_only = st.checkbox("Younger than target", value=False)
            with f8:
                cheaper_only = st.checkbox("Cheaper than target", value=False)
            with f9:
                undervalued_only = st.checkbox("Only undervalued", value=False)

            filtered_df = similar_pool_df.copy()
            filtered_df = filtered_df[pd.to_numeric(filtered_df["age_at_valuation"], errors="coerce") <= max_age]
            filtered_df = filtered_df[pd.to_numeric(filtered_df["target_market_value"], errors="coerce") <= max_value]
            filtered_df = filtered_df[filtered_df[score_col] >= (min_overall_fit / 100)]
            if role_score_col:
                filtered_df = filtered_df[filtered_df[role_score_col] >= (min_role_fit / 100)]
            if younger_only:
                filtered_df = filtered_df[filtered_df["is_younger"].fillna(False).astype(bool)].copy()
            if cheaper_only:
                filtered_df = filtered_df[filtered_df["is_cheaper"].fillna(False).astype(bool)].copy()
            if undervalued_only:
                filtered_df = filtered_df[filtered_df["is_undervalued"].fillna(False).astype(bool)].copy()

            filtered_df = filtered_df.sort_values(by=score_col, ascending=False).head(max_results)

            st.caption(f"Showing {len(filtered_df):,} of {len(similar_pool_df):,} role-aware candidate matches after recruitment filters.")

            if filtered_df.empty:
                st.warning("No candidates match the current recruitment filters. Relax age, value, or fit filters to broaden the shortlist.")
            else:
                # Format for display using the new position-aware function
                target_position = target_data["position_group_raw"]
                final_display = build_similarity_display_table(
                    res_df=filtered_df,
                    target_position=target_position,
                    mode=actual_mode
                )

                st.dataframe(
                    final_display,
                    width="stretch",
                    hide_index=True,
                    column_config=similarity_column_config(),
                )
                with st.expander("How to read this table"):
                    st.markdown(
                        "- **Club / League** uses current Transfermarkt club and competition context when available; enriched stats still follow the displayed stats season.\n"
                        "- **Stat Score** measures performance-profile similarity before role weighting.\n"
                        "- **Role Score** measures tactical compatibility. 100% means the same role; lower scores indicate compatible adjacent roles.\n"
                        "- **Final Score** combines statistical similarity, role compatibility, and preferred-foot fit.\n"
                        "- **Price Status** compares the candidate's current value with the target player using a ±15% similar-cost band."
                    )
                    if actual_mode == "enriched":
                        st.markdown("- **Enriched mode** uses Kaggle Top 5 League advanced stats and includes only accepted external matches.")
                    elif target_position == "Goalkeeper":
                        st.markdown("- **Basic goalkeeper mode** relies mainly on age, height, and playing time because detailed save metrics are not available.")
                    elif target_position == "Defender":
                        st.markdown("- **Basic defender mode** uses available profile and discipline metrics; detailed defensive actions are only available in Enriched mode.")
                    else:
                        st.markdown("- **Basic attacking mode** uses core contribution metrics: Goals/90, Assists/90, and Cards/90.")

                # Download
                if _OPENPYXL_AVAILABLE:
                    _sim_excel = build_similarity_excel(filtered_df, target_name=target_player_name)
                    st.download_button(
                        label=f"💾 Download Comparison for {target_player_name} (Excel)",
                        data=_sim_excel,
                        file_name=f"alternatives_{safe_filename_label(target_player_name)}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                else:
                    st.warning("openpyxl is not installed  -  Excel export unavailable. Install it with `pip install openpyxl`.")
                    st.download_button(
                        label=f"💾 Download Alternatives for {target_player_name} (CSV fallback)",
                        data=add_download_context_columns(filtered_df).to_csv(index=False),
                        file_name=f"alternatives_{safe_filename_label(target_player_name)}.csv",
                        mime="text/csv",
                    )
        else:
            st.info("No similar players found for the selected role mode. Try Compatible Roles or Broad Position Group.")

        st.divider()
        methodology_box("Similarity search supports statistical and role layers. **Enriched Mode** uses advanced Top 5 League metrics matched through season-safe enrichment. **Role-Aware Mode** adds Transfermarkt role metadata, compatible-role filtering, and a small foot/side-fit bonus. Market value and predicted valuation are used strictly for post-match filtering and comparison, never for calculating similarity.")
        disclaimer_box("Candidates identified here are statistical leads based on predictive modeling of recent performance and should be interpreted as a starting point for professional human scouting.")


# =========================
# TAB 3: SCOUTING RATIONALE
# =========================
with tab_rationale:
    st.markdown("### 🧠 Why This Player?")
    st.write("Plain-language scouting notes that explain why a player deserves review.")
    st.info(
        "This page is for scouting decisions, not technical model debugging. "
        "It uses the same player data, value estimate, age, minutes, and role context shown in the shortlist."
    )

    pool_mode = st.radio(
        "Search in:",
        ["Current Shortlist", "All Modelled Players"],
        horizontal=True,
        help=(
            "Current Shortlist uses the active filtered shortlist from the first tab. "
            "All Modelled Players searches the full prediction output."
        ),
    )

    if pool_mode == "Current Shortlist":
        pool_df = st.session_state.get("active_candidate_shortlist", load_scouting_data()).copy()
    else:
        pool_df = load_all_predictions()
        if pool_df.empty:
            pool_df = load_scouting_data()

    if pool_df.empty:
        st.warning("No players available in the selected pool.")
    else:
        name_col = "name" if "name" in pool_df.columns else "Player Name"
        pool_df = (
            pool_df.dropna(subset=[name_col])
            .drop_duplicates(subset=["player_id"] if "player_id" in pool_df.columns else [name_col])
            .copy()
        )
        if pool_mode == "All Modelled Players":
            pool_df = pool_df.sort_values(by=name_col)

        st.caption(
            f"{len(pool_df):,} players available · "
            f"{'follows the active filters in Find Undervalued Players' if pool_mode == 'Current Shortlist' else 'full modelled player pool'}"
        )

        selected_player_name = st.selectbox(
            "Select player to review",
            pool_df[name_col].tolist(),
            key="player_rationale_select",
        )
        selected_row = pool_df[pool_df[name_col] == selected_player_name].iloc[0]

        # ── Compute confidence and action once (used in hero + action card) ──
        confidence = compute_confidence_level(selected_row)
        confidence_color = {"High": "🟢", "Medium": "🟡", "Low": "🔴"}.get(confidence, "⚪")

        gap_val = _safe_float(selected_row.get("undervalued_pct", 0))
        contract_st = selected_row.get("contract_status", "Unknown")
        age_val = _safe_float(selected_row.get("age_at_valuation", 30), default=30.0)
        if gap_val >= 1.0 and contract_st in ("Expiring", "Expired"):
            action = "🔥 Prioritize  -  Expiring Contract"
        elif gap_val >= 1.0 and age_val <= 23:
            action = "⭐ Prioritize  -  Young & Undervalued"
        elif gap_val >= 1.0:
            action = "✅ Prioritize Review"
        elif gap_val >= 0.5:
            action = "👀 Monitor"
        else:
            action = "📋 Needs Review"

        # ── Hero: name + tier + contract badges + confidence ─────────────────
        hero_comp_id = selected_row.get("current_club_domestic_competition_id", None)
        hero_contract = selected_row.get("contract_status", "Unknown")
        badges_html = (
            league_tier_badge(hero_comp_id)
            + " "
            + contract_status_badge(hero_contract)
            + f' <span class="hero-meta" style="margin-left:6px;">{confidence_color} Confidence: {html.escape(confidence)}</span>'
        )
        hero_meta = [
            str(selected_row.get("position_group_raw", "N/A")),
            f"{age_val:.1f} yrs",
            f"{_safe_float(selected_row.get('minutes_last_season', 0)):,.0f} min",
            current_club_display_label(selected_row),
        ]
        render_player_hero(selected_player_name, badges_html=badges_html, meta_items=hero_meta)

        # ── Primary KPIs: 4 cards that matter most for scouting ──────────────
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            kpi_card("Current Value", format_currency(selected_row.get("target_market_value")))
        with c2:
            kpi_card("Estimated Value", format_currency(selected_row.get("predicted_value")))
        with c3:
            kpi_card("Value Gap", format_percentage(selected_row.get("undervalued_pct", 0)))
        with c4:
            kpi_card("Suggested Action", action)

        # ── Form strip: single row replaces 4 form KPI cards ─────────────────
        form_val = selected_row.get("form_trend", "Insufficient Data")
        render_form_strip(
            form_val=str(form_val),
            goals_delta=selected_row.get("form_trend_goals", None),
            assists_delta=selected_row.get("form_trend_assists", None),
            mins_ratio=selected_row.get("form_trend_minutes", None),
        )

        evidence_df = build_scouting_signals(selected_row)

        # Phase 10: personalized rationale with confidence level
        rationale_text = build_rationale_summary(selected_player_name, evidence_df, row=selected_row)
        st.markdown(rationale_text)
        st.caption(f"{confidence_color} **Data Confidence: {confidence}**  -  based on league tier, minutes played, contract availability, and advanced stats coverage.")

        scout_col_1, scout_col_2 = st.columns(2)
        with scout_col_1:
            st.markdown("##### 📈 Key Scouting Signals")
            if evidence_df.empty:
                st.write("No scouting evidence available for this player.")
            else:
                for _, signal in evidence_df.iterrows():
                    st.success(f"{signal['Signal']}: {signal['Scout Interpretation']}")

        with scout_col_2:
            st.markdown("##### ⚠️ Scout Checks Before Action")
            for check in build_scout_checks(selected_row):
                st.warning(check)
            # Phase 8: add league tier contextual note
            tier_note = get_scout_note(selected_row.get("current_club_domestic_competition_id", None))
            if tier_note:
                tier_short = get_tier_short(selected_row.get("current_club_domestic_competition_id", None))
                st.info(f"🌍 **League Tier ({tier_short}):** {tier_note}")
            # Phase 9: add contract contextual note
            contract_status = selected_row.get("contract_status", "Unknown")
            contract_note = get_contract_scout_note(contract_status)
            if contract_note:
                contract_badge_text = get_contract_badge(contract_status)
                st.info(f"📋 **Contract ({contract_badge_text}):** {contract_note}")

        st.markdown("##### 📊 Evidence Table")
        if evidence_df.empty:
            st.info("No evidence table is available for this player.")
        else:
            st.dataframe(evidence_df, width="stretch", hide_index=True)

        # ── Phase 12: PDF Scouting Brief ─────────────────────────────────────
        st.divider()
        st.markdown("##### 📄 Export Scouting Brief")
        if not _FPDF_AVAILABLE:
            st.warning("fpdf2 is not installed  -  PDF export unavailable. Install it with `pip install fpdf2`.")
        else:
            if st.button("📄 Generate Scouting Brief", key="btn_generate_pdf"):
                _scout_checks_list = build_scout_checks(selected_row)
                # Add league tier and contract notes to checks if present
                _tier_note = get_scout_note(selected_row.get("current_club_domestic_competition_id", None))
                if _tier_note:
                    _tier_short_label = get_tier_short(selected_row.get("current_club_domestic_competition_id", None))
                    _scout_checks_list = [f"League Tier ({_tier_short_label}): {_tier_note}"] + _scout_checks_list
                _contract_note = get_contract_scout_note(selected_row.get("contract_status", "Unknown"))
                if _contract_note:
                    _contract_badge_label = get_contract_badge(selected_row.get("contract_status", "Unknown"))
                    _scout_checks_list = _scout_checks_list + [f"Contract ({_contract_badge_label}): {_contract_note}"]

                _pdf_bytes = build_scouting_pdf(
                    player_name=selected_player_name,
                    row=selected_row,
                    rationale_text=rationale_text,
                    signals_df=evidence_df,
                    scout_checks=_scout_checks_list,
                )
                if _pdf_bytes:
                    st.download_button(
                        label=f"⬇️ Download PDF  -  {selected_player_name}",
                        data=_pdf_bytes,
                        file_name=f"scouting_brief_{safe_filename_label(selected_player_name)}.pdf",
                        mime="application/pdf",
                        key="btn_download_pdf",
                    )
                else:
                    st.error("PDF generation failed. Check that fpdf2 is correctly installed.")


# =========================
# TAB 4: BACKTEST RESULTS
# =========================
with tab_backtest:
    st.markdown("### 📈 Backtest Results")
    st.write("Historical outcome validation for scouting leads flagged by the fixed performance-only model.")

    backtest_outputs = load_temporal_backtest_outputs()
    summary_df = backtest_outputs["summary"].copy()
    candidates_df = backtest_outputs["candidates"].copy()
    by_gap_df = backtest_outputs["by_value_gap"].copy()
    by_position_df = backtest_outputs["by_position"].copy()
    report = backtest_outputs["report"]

    if summary_df.empty or candidates_df.empty:
        st.info("Phase 7 outputs are not available yet. Run `python src/temporal_backtest.py` and `python src/validate_temporal_backtest.py`.")
        if backtest_outputs["missing"]:
            st.caption("Missing outputs: " + ", ".join(backtest_outputs["missing"]))
    else:
        status = report.get("status", "UNKNOWN")
        if status == "PASS":
            st.success("Temporal validation status: PASS")
        elif status == "WARN":
            st.warning("Temporal validation status: WARN. Review `outputs/temporal_validation/temporal_validation_report.json`.")
        else:
            st.error("Temporal validation status: FAIL or not validated. Review `outputs/temporal_validation/temporal_validation_report.json`.")

        summary_df["snapshot_date"] = summary_df["snapshot_date"].astype(str)
        lead_overall = summary_df[
            (summary_df["population"] == "Scouting Leads") & (summary_df["snapshot_date"] == "ALL")
        ]
        baseline_overall = summary_df[
            (summary_df["population"] == "Eligible Baseline") & (summary_df["snapshot_date"] == "ALL")
        ]
        lead_row = lead_overall.iloc[0] if not lead_overall.empty else pd.Series(dtype="object")
        baseline_row = baseline_overall.iloc[0] if not baseline_overall.empty else pd.Series(dtype="object")

        hit_12m_lift = lead_row.get("hit_rate_12m_25pct", np.nan) - baseline_row.get("hit_rate_12m_25pct", np.nan)

        b1, b2, b3, b4, b5 = st.columns(5)
        with b1:
            kpi_card("Historical Leads", f"{int(lead_row.get('row_count', 0)):,}")
        with b2:
            kpi_card("6M Hit Rate", format_percentage(lead_row.get("hit_rate_6m_25pct", np.nan)))
        with b3:
            kpi_card("12M Hit Rate", format_percentage(lead_row.get("hit_rate_12m_25pct", np.nan)))
        with b4:
            kpi_card("Median 12M Growth", format_percentage(lead_row.get("median_growth_12m_pct", np.nan)))
        with b5:
            kpi_card("12M Lift vs Baseline", format_percentage_points(hit_12m_lift))

        methodology_box(
            "This is a fixed-model retrospective signal audit. It replays historical snapshots with the existing performance-only model and checks future market-value movement after 6 and 12 months."
        )

        chart_col_1, chart_col_2 = st.columns(2)
        with chart_col_1:
            gap_chart = by_gap_df[by_gap_df["population"] == "Scouting Leads"].copy()
            if not gap_chart.empty and "hit_rate_12m_25pct" in gap_chart.columns:
                gap_chart["12M Hit Rate"] = pd.to_numeric(gap_chart["hit_rate_12m_25pct"], errors="coerce") * 100
                fig_gap = px.bar(
                    gap_chart.dropna(subset=["12M Hit Rate"]),
                    x="value_gap_bucket",
                    y="12M Hit Rate",
                    template=PLOTLY_TEMPLATE,
                    labels={"value_gap_bucket": "Value Gap Bucket", "12M Hit Rate": "12M Hit Rate (%)"},
                    title="12M Hit Rate by Value Gap",
                    color_discrete_sequence=["#0f766e"],
                )
                fig_gap.update_layout(height=390, title_x=0.02, xaxis_tickangle=-20)
                st.plotly_chart(fig_gap, width="stretch")
        with chart_col_2:
            position_chart = by_position_df[by_position_df["population"] == "Scouting Leads"].copy()
            if not position_chart.empty and "median_growth_12m_pct" in position_chart.columns:
                position_chart["Median 12M Growth"] = pd.to_numeric(position_chart["median_growth_12m_pct"], errors="coerce") * 100
                fig_position = px.bar(
                    position_chart.dropna(subset=["Median 12M Growth"]),
                    x="position_group_raw",
                    y="Median 12M Growth",
                    template=PLOTLY_TEMPLATE,
                    labels={"position_group_raw": "Position", "Median 12M Growth": "Median 12M Growth (%)"},
                    title="Median 12M Growth by Position",
                    color_discrete_sequence=["#2563eb"],
                )
                fig_position.update_layout(height=390, title_x=0.02)
                st.plotly_chart(fig_position, width="stretch")

        growth_df = candidates_df.copy()
        if "growth_12m_pct" in growth_df.columns:
            growth_df["12M Growth"] = pd.to_numeric(growth_df["growth_12m_pct"], errors="coerce") * 100
            growth_df = growth_df.dropna(subset=["12M Growth"])
            if not growth_df.empty:
                fig_growth = px.histogram(
                    growth_df,
                    x="12M Growth",
                    nbins=40,
                    template=PLOTLY_TEMPLATE,
                    labels={"12M Growth": "12M Growth (%)"},
                    title="Future Value Growth Distribution",
                    color_discrete_sequence=["#f59e0b"],
                )
                fig_growth.add_vline(x=25, line_dash="dash", line_color=REFERENCE_LINE_COLOR)
                fig_growth.update_layout(height=360, title_x=0.02, yaxis_title="Historical Leads")
                st.plotly_chart(fig_growth, width="stretch")

        st.markdown("#### Historical Scouting Leads")
        table_source = candidates_df.copy()
        if "snapshot_date" in table_source.columns:
            snapshots = ["All"] + sorted(table_source["snapshot_date"].dropna().astype(str).unique().tolist())
            selected_snapshot = st.selectbox("Snapshot", snapshots, index=0)
            if selected_snapshot != "All":
                table_source = table_source[table_source["snapshot_date"].astype(str) == selected_snapshot].copy()

        if "growth_12m_pct" in table_source.columns:
            table_source["growth_12m_pct"] = pd.to_numeric(table_source["growth_12m_pct"], errors="coerce")
            table_source = table_source.sort_values("growth_12m_pct", ascending=False, na_position="last")

        display_backtest = build_backtest_display_table(table_source.head(300))
        st.dataframe(
            display_backtest,
            width="stretch",
            hide_index=True,
            column_config=backtest_column_config(),
        )
